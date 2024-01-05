#T23_11_v2
# Створення робочої книги MS Excel
# Зображення графіку декількох функцій на інтервалі [a,b] у n точках

from openpyxl import *
from openpyxl.chart import (
    ScatterChart,
    LineChart,
    Reference,
    Series,
)

from T20.t20_11_tabulate_v3 import *
from math import sin


def plotfunc2(a, b, n, *f):
    '''Зображує графік функцій *f на інтервалі [a,b] у n точках
    '''
    wb = Workbook()             # створити робочу книгу 
    ws = wb.active              # вибрати активний робочий аркуш

    # створити графік
    chart1 = ScatterChart()
    chart1.legend = None

    tabb = False                # чи було табульовано першу функцію
    k = 1                       # номер стовпчика з даними
    for ff in f:
        k += 1
        if not tabb:
            x, y = tabulate(ff, a, b, n) #табулювати функцію
            tabb = True
            ws.append(['x', ff.__name__])# додати заголовки стовпчиків
            for i in range(n):
                ws.append([x[i], y[i]]) # додати дані
            xdata = Reference(ws, min_col=1, min_row=2, max_row=n+1)
        else:
            y = gety(ff, x)             #отримати y за x для функції
            # додати заголовок стовпчика
            ws.cell(row = 1, column = k, value = ff.__name__)
            for i in range(n):
                ws.cell(row = i + 2, column = k, value = y[i]) # додати дані
        # додати стовпчик до графіка
        ydata = Reference(ws, min_col=k, min_row=2, max_row=n+1)
        s = Series(ydata, xvalues=xdata)
        chart1.append(s)

    ws.add_chart(chart1, "E1") # додати графік
    wb.save('graphics.xlsx')   #зберегти робочу книгу

if __name__ == '__main__':
    n = int(input('Кількість точок: '))
    a = float(input('Початок відрізку: '))
    b = float(input('Кінець відрізку: '))

    plotfunc2(a, b, n, fun, sin)
    
