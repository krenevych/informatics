#from openpyxl import load_workbook

import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("sample1.xlsx")

#wb = load_workbook("sample1.xlsx")

print(wb.get_sheet_names())

# grab the active worksheet
ws = wb.active

ws['A1'] = 10

ws['B1'] = 20


ws['C1'] = '= A1 + B1'

ws['D1'] = '=C1'

c = ws.cell(row = 2, column = 1)
c.value = 11111

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
#wb.save("sample1.xlsx")

wb.save("sample2.xlsx")
