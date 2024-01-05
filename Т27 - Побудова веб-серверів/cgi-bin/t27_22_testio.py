#T27_22 Класи введення-виведення тестів.

from abc import ABCMeta, abstractmethod
from openpyxl import *
from t27_21_quiz import *

class TestIO(metaclass = ABCMeta):
    """Абстрактний клас введення тестів та виведення результатів.

        self.quissuite - тести (об'єкт класу QuisSuite)
        self.urn - розташування ресурсу з тестами (файл або база даних)
        self.params - додаткові параметри для читання тестів
        self.users - список користувачів - кортежів (користувач, пароль)
        self.results = список результатів - кортежів Result
    """

    def __init__(self, quissuite, urn, **params):
        self.quissuite = quissuite
        self.urn = urn
        self.params = params
        self._users = []
    
    @abstractmethod
    def read(self):
        """Читати тести.
           Повертає список об'єктів класу Quiz та список результатів"""
        pass

    @property
    @abstractmethod
    def users(self):
        "Повертає список кортежів (користувач, пароль)"
        pass

    @abstractmethod
    def writeresult(self, result):
        "Писати результат тесту з result."
        pass


class TestExcelIO(TestIO):
    """Клас введення тестів з файлу MS Excel та виведення результатів.

        self.quissuite - тести (об'єкт класу QuisSuite)
        self.urn - розташування ресурсу з тестами (файл MS Excel)
        self.wb - робоча книга MS Ecxel з тестами
        self.resurn - розташування ресурсу з результатами (текстовий файл)
        self.users - список користувачів - списків [користувач, пароль]
    """

    def __init__(self, quissuite, urn):
        TestIO.__init__(self, quissuite, urn)
        self.wb = load_workbook(urn)
        self.resurn = urn + '.txt'
        # прочитати користувачів
        self._users = self._readws("$users")
#        print(self._users)

    def _readws(self, ws_name):
        "Повертає список списків даних робочого аркушу ws_name."
        cnt = []
        ws = self.wb[ws_name]
        for r in range(ws.min_row, ws.max_row + 1):
            record = []
            for c in range(ws.min_column, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v:
                    record.append(v)
            if record:
                cnt.append(record)
        return cnt
            
        
    def read(self):
        """Читати тести.
           Повертає список об'єктів класу Quiz та список результатів"""
        quizzes = []
        for ws in self.wb:
            # якщо не спеціальний аркуш з користувачами або результатами
            if not ws.title.startswith("$"):
                quizzes.append(self._readquiz(ws))
        return quizzes
                
    def _readquiz(self, ws):
        "Повертає тест з робочого аркушу ws."
        theme = ws.cell(row=1, column=1).value
        quiz = Quiz(self.quissuite, theme)
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v:   # якщо питання
                typ = ws.cell(row=r, column=3).value
                points = ws.cell(row=r, column=4).value
                quest = Question(quiz, v, typ, points)
                quiz.questions.append(quest)
            else:   # якщо відповідь
                atext = ws.cell(row=r, column=3).value
                avalue = bool(ws.cell(row=r, column=4).value)
                quest.answers.append(Answer(quest, atext, avalue))
        return quiz

    @property
    def users(self):
        "Повертає список кортежів (користувач, пароль)"
        return self._users


    def writeresult(self, result):
        "Писати результат тесту з result."
        with open(self.resurn, 'a') as f:
            f.write(str(result) + '\n')
            
if __name__ == '__main__':
    filename = input("файл: ")
    suite = QuizSuite(TestExcelIO, filename)
    shift = 0
    for quiz in suite.quizzes:
        print(shift*' ', quiz)
        shift += 4
        for quest in quiz.questions:
            print(shift*' ', quest)
            shift += 4
            for ans in quest.answers:
                print(shift*' ', ans)
            shift -= 4
        shift -= 4
    print(suite.getthemes())
        


        

